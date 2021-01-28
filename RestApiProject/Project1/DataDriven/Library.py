import json
import jsonpath
import requests
import openpyxl


class Common:
    def __init__(self, Path, Sheetname):
        global workbook, sheet
        # open excel sheet
        workbook = openpyxl.load_workbook(Path, Sheetname)
        # Read Sheet 1
        sheet = workbook[Sheetname]

    def fetch_row_count(self):
        # Fetching Number of rows
        rows = sheet.max_row
        return rows

    def fetch_column_count(self):
        column = sheet.max_column
        return column

    def fetch_key_name(self):
        colmn = sheet.max_column
        li = []
        for i in range(1, colmn + 1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range(1, c + 1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest
        workbook.close()