import openpyxl
import requests
import json
import jsonpath


def test_add_multiple_student_data():
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    # open the student json file in read mode
    file = open("D:\\RestApiProject\\ApiDataDrivenTestCases\\addstudent.json", 'r')
    json_file = file.read()
    json_request = json.loads(json_file)
    # open excel sheet
    workbook = openpyxl.load_workbook("D:\\RestApiProject\\ApiDataDrivenTestCases\\test123.xlsx")
    # Read Sheet 1
    sheet = workbook['Sheet1']
    # Fetching Number of rows
    rows = sheet.max_row
    print(rows)
    # Reading the data form excel sheet
    for i in range(2, rows+1):
        firstname = sheet.cell(row=i, column=1)
        print(firstname)
        middlename = sheet.cell(row=i, column=2)
        lastname = sheet.cell(row=i, column=3)
        dateofbirth = sheet.cell(row=i, column=4)
        json_request['first_name'] = firstname.value
        json_request['middle_name'] = middlename.value
        json_request['last_name'] = lastname.value
        json_request['date_of_birth'] = dateofbirth.value
        response = requests.post(api_url, json_request)
        print(response)
        print(response.status_code)
        #assert response.status_code == 201

    workbook.close()
